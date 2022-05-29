#1. reading file from excel

import openpyxl
path = (r"C:/Users/navan/OneDrive/Desktop/SaltBakerySales.xlsx")
workbook = openpyxl.load_workbook(path)
# if you many sheets in exel file
#sheet = workbook.active
sheet = workbook.get_sheet_by_name("Sales") # sheet  name

row = (sheet.max_row)
col = (sheet.max_column)

for r in range(1,row+1):
    for c in range(1,col+1):
        print(sheet.cell(row=r,column=c).value, end=" ")
    print()
